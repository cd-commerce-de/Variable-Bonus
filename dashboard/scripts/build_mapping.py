import json
from openpyxl import load_workbook

wb = load_workbook('/mnt/user-data/uploads/TOC__Brands_and_Products_List_for_Reports_.xlsx', data_only=True)
ws = wb['ASIN Report']

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(headers) if h}

mapping = {}
for r in range(2, ws.max_row + 1):
    asin = ws.cell(row=r, column=col['ASIN']).value
    if not asin:
        continue
    asin = asin.strip()
    if asin in mapping:
        continue  # keep first occurrence (verified: no conflicting duplicates)
    mapping[asin] = {
        'brand': ws.cell(row=r, column=col['Brand']).value,
        'stage': ws.cell(row=r, column=col['Stage']).value,
        'status': ws.cell(row=r, column=col['Status']).value,
        'product': ws.cell(row=r, column=col['Product']).value,
        'product_code': ws.cell(row=r, column=col['Product Code']).value,
        'source': ws.cell(row=r, column=col['Source']).value,
        'title': ws.cell(row=r, column=col['Title']).value,
    }

with open('/home/claude/dashboard/mapping/toc_mapping.json', 'w', encoding='utf-8') as f:
    json.dump(mapping, f, ensure_ascii=False, indent=1)

print('mapped ASINs:', len(mapping))
stages = {}
for v in mapping.values():
    stages[v['stage']] = stages.get(v['stage'], 0) + 1
print('by stage:', stages)
