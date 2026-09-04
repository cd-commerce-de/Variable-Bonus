"""
Pulls Q3 (Jul-Sep) Green/Gold revenue + margin targets, the manual gate
value, and all Finance-set rates/weights out of the Variable Bonus
Calculator workbook, then divides revenue targets by 3 to get a monthly
(August) figure.

This is the interim approach explicitly requested: real quarterly targets
from the workbook, evenly divided by 3, rather than the weekly-sourced
decomposition used for the calculator's own Actuals. When real monthly
targets are set in the workbook (per the earlier monthly restructuring of
the All Tracks tab), swap this for reading those columns directly instead
of dividing.

Usage: python3 scripts/extract_targets.py <calculator.xlsx> [--quarter Q3]
Writes: mapping/targets.json
"""
import json
import argparse
from openpyxl import load_workbook

# Q3 (Jul/Aug/Sep) quarter-total block in the ORIGINAL (unrestructured)
# All Tracks layout: AE=GreenTargetRev AF=GoldTargetRev AG=ActualRev
# AH=GreenMargin% AI=GoldMargin% AJ=ActualMargin% AK=QGauto AL=QGgate(manual)
QUARTER_COLS = {
    'Q1': {'green_rev': 5, 'gold_rev': 6, 'green_margin': 8, 'gold_margin': 9, 'gate': 12},
    'Q2': {'green_rev': 17, 'gold_rev': 18, 'green_margin': 20, 'gold_margin': 21, 'gate': 24},
    'Q3': {'green_rev': 31, 'gold_rev': 32, 'green_margin': 34, 'gold_margin': 35, 'gate': 38},
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx_path')
    ap.add_argument('--quarter', default='Q3')
    ap.add_argument('--out', default='mapping/targets.json')
    args = ap.parse_args()

    wb = load_workbook(args.xlsx_path, data_only=True)
    ws = wb['📋 All Tracks']
    cfg = wb['⚙️ Config']
    cols = QUARTER_COLS[args.quarter]

    def cell(r, c):
        return ws.cell(row=r, column=c).value

    def row_target(r, label=None):
        gr, gd = cell(r, cols['green_rev']), cell(r, cols['gold_rev'])
        gm, dm = cell(r, cols['green_margin']), cell(r, cols['gold_margin'])
        gate = cell(r, cols['gate'])
        return {
            'label': label,
            'quarter_green_rev': gr, 'quarter_gold_rev': gd,
            'monthly_green_rev': (gr / 3) if isinstance(gr, (int, float)) else None,
            'monthly_gold_rev': (gd / 3) if isinstance(gd, (int, float)) else None,
            'green_margin_pct': gm, 'gold_margin_pct': dm,
            'gate': gate,
        }

    # ---- R&D: per named product (row -> product code parsed from Label) ----
    rd_rows = {
        7: 'GWK', 8: 'DSD', 9: 'TSE', 10: 'TSF', 11: 'KMS', 12: 'FES',
        13: 'FRM02', 14: 'WHS', 15: 'SLP', 16: 'VZW', 17: 'AKS', 18: 'WGH25', 19: 'KMK',
    }
    rd_targets = {}
    for r, code in rd_rows.items():
        label = cell(r, 3)  # column C
        rd_targets[code] = row_target(r, label)

    # ---- Launch Manager: Germany (23) / PAN EU (24) ----
    launch_targets = {
        'germany': row_target(23, 'Germany F3M'),
        'pan_eu': row_target(24, 'PAN EU Expansion F3M'),
    }

    # ---- Brand Manager: brand -> {PY1, Y1, Discontinued} ----
    bm_blocks = {
        'Tarpofix': (28, 29, 30), 'Darwin': (31, 32, 33), 'Planenfux': (34, 35, 36),
        'Heimfleiss': (38, 39, 40), 'Mattenheld': (41, 42, 43), 'PD': (45, 46, 47),
        'Nasswerk': (49, 50, 51), 'PoolLöwe': (52, 53, 54), 'TeichHeld': (55, 56, 57),
    }
    stage_names = ['PY1', 'Y1 (F4-12)', 'Discontinued']
    bm_targets = {}
    for brand, (r_py1, r_y1, r_disc) in bm_blocks.items():
        bm_targets[brand] = {
            stage_names[0]: row_target(r_py1),
            stage_names[1]: row_target(r_y1),
            stage_names[2]: row_target(r_disc),
        }

    # ---- Marketplace (row 64) ----
    marketplace_target = row_target(64, 'Marketplace')

    # ---- Config: rates & weights ----
    rates = {
        'rd_team': {'green': cfg['C10'].value, 'gold': cfg['D10'].value, 'team_size': cfg['C5'].value},
        'brand_manager': {'green': cfg['C11'].value, 'gold': cfg['D11'].value},
        'launch_mgr_germany': {'green': cfg['C12'].value, 'gold': cfg['D12'].value},
        'launch_mgr_pan_eu': {'green': cfg['C13'].value, 'gold': cfg['D13'].value},
        'marketplace': {'green': cfg['C14'].value, 'gold': cfg['D14'].value, 'team_size': cfg['C6'].value},
    }
    stage_weights = {
        'PY1': {'weight': cfg['C18'].value, 'eff_green': cfg['D18'].value, 'eff_gold': cfg['E18'].value},
        'Y1 (F4-12)': {'weight': cfg['C19'].value, 'eff_green': cfg['D19'].value, 'eff_gold': cfg['E19'].value},
        'Discontinued': {'weight': cfg['C20'].value, 'eff_green': cfg['D20'].value, 'eff_gold': cfg['E20'].value},
    }

    result = {
        'source_quarter': args.quarter,
        'note': 'Monthly targets = quarterly target / 3 (interim, per instruction). Replace with real monthly targets once available.',
        'rd_team': rd_targets,
        'launch_manager': launch_targets,
        'brand_manager': bm_targets,
        'marketplace': marketplace_target,
        'rates': rates,
        'stage_weights': stage_weights,
    }
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=1, ensure_ascii=False)
    print('Wrote', args.out)


if __name__ == '__main__':
    main()
