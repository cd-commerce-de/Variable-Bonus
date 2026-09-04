import json
import sys
from datetime import datetime
from openpyxl import load_workbook

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/TOC__Brands_and_Products_List_for_Reports_.xlsx'

wb = load_workbook(xlsx_path, data_only=True)
ws = wb['ASIN Report']

headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
col = {h: i + 1 for i, h in enumerate(headers) if h}


def iso_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return None


mapping = {}
for r in range(2, ws.max_row + 1):
    asin = ws.cell(row=r, column=col['ASIN']).value
    if not asin:
        continue
    asin = asin.strip()
    if asin in mapping:
        continue  # keep first occurrence (verified: no conflicting duplicates)
    mapping[asin] = {
        'brand': ws.cell(row=r, column=col['Brand']).value,
        'toc_stage_snapshot': ws.cell(row=r, column=col['Stage']).value,  # reference only -- the dashboard computes the real, month-specific stage from the dates below
        'status': ws.cell(row=r, column=col['Status']).value,
        'product': ws.cell(row=r, column=col['Product']).value,
        'product_code': ws.cell(row=r, column=col['Product Code']).value,
        'source': ws.cell(row=r, column=col['Source']).value,
        'title': ws.cell(row=r, column=col['Title']).value,
        'launch_date': iso_date(ws.cell(row=r, column=col['Launch Date']).value),
        'discontinued_start_date': iso_date(ws.cell(row=r, column=col['Discontinued Start Date']).value),
        'quality_issue_start_date': iso_date(ws.cell(row=r, column=col['Quality Issue Start Date']).value),
    }

with open('/home/claude/dashboard/mapping/toc_mapping.json', 'w', encoding='utf-8') as f:
    json.dump(mapping, f, ensure_ascii=False, indent=1)

print('mapped ASINs:', len(mapping))
no_launch_date = sum(1 for v in mapping.values() if not v['launch_date'])
print(f'ASINs with no Launch Date (stage cannot be computed, will fall back to toc_stage_snapshot): {no_launch_date}')
snapshot_stages = {}
for v in mapping.values():
    snapshot_stages[v['toc_stage_snapshot']] = snapshot_stages.get(v['toc_stage_snapshot'], 0) + 1
print('by TOC snapshot stage (reference only):', snapshot_stages)

