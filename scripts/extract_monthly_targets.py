"""
Extracts real per-month Good/Better/Best (Green/Gold-equivalent... here
3-tier) revenue + profit-margin targets, sourced directly from:

  - BM Scorecard 3  -> Launch Manager (LM (F3M), Expansion (F3M)) and
                       Brand Manager (per brand, per stage: PY1/Y1/Discontinued)
  - Leadership Scorecard 3, starting row 13 -> R&D (revenue only; margin
    left blank, as instructed -- no margin target exists for R&D)

Replaces the earlier "quarterly / 3" interim approach with real monthly
figures for whichever month is requested.

Usage: python3 scripts/extract_monthly_targets.py <calculator.xlsx> --month 2026-08
Writes: mapping/targets_monthly.json
"""
import json
import re
import argparse
from datetime import datetime
from openpyxl import load_workbook


def find_month_columns(ws, header_row=3):
    """Scan a header row for GOOD/BETTER/BEST triples and figure out which
    calendar month each belongs to, using the nearest date cell that
    follows. Returns {month_number: {'good': col, 'better': col, 'best': col}}."""
    max_col = ws.max_column
    triples = []
    c = 1
    while c <= max_col - 2:
        v = ws.cell(row=header_row, column=c).value
        if v == 'GOOD' and ws.cell(row=header_row, column=c + 1).value == 'BETTER' and ws.cell(row=header_row, column=c + 2).value == 'BEST':
            triples.append(c)
            c += 3
        else:
            c += 1

    result = {}
    for good_col in triples:
        month_num = None
        for c2 in range(good_col, min(good_col + 20, max_col + 1)):
            v = ws.cell(row=header_row, column=c2).value
            if isinstance(v, datetime):
                month_num = v.month
                break
        if month_num:
            result[month_num] = {'good': good_col, 'better': good_col + 1, 'best': good_col + 2}
    return result


def find_section(ws, col_a_pattern, start_row=1, end_row=None):
    """Yield (row, matched_text) for column-A cells matching a regex."""
    end_row = end_row or ws.max_row
    for r in range(start_row, end_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and re.match(col_a_pattern, str(v)):
            yield r, v


def gbb(ws, row, cols):
    return {
        'good': ws.cell(row=row, column=cols['good']).value,
        'better': ws.cell(row=row, column=cols['better']).value,
        'best': ws.cell(row=row, column=cols['best']).value,
    }


def extract_bm_scorecard(ws, cols):
    """LM(F3M)/Expansion(F3M) + per-brand PY1/Y1/Discontinued sections."""
    # Row map: column-A label -> row number, in document order
    labels = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            labels[r] = str(v).strip()
        if v == 'Total (Amazon EU)':
            break  # stop before grand-total sections, not needed

    rows_sorted = sorted(labels.keys())

    # ---- Launch Manager ----
    def find_after(header_row, contains):
        for r in rows_sorted:
            if r <= header_row:
                continue
            if labels[r].startswith(('BM', 'LM', 'Expansion')) and r != header_row:
                # hit the next section header before finding it
                if re.match(r'^(LM|Expansion|BM\d)', labels[r]):
                    break
            if contains.lower() in labels[r].lower():
                return r
        return None

    lm_header = next(r for r in rows_sorted if labels[r] == 'LM (F3M)')
    exp_header = next(r for r in rows_sorted if labels[r] == 'Expansion (F3M)')
    lm_rev_row = find_after(lm_header, 'Rolling F3M Revenue')
    lm_profit_row = find_after(lm_header, 'Rolling F3M Profit')
    exp_rev_row = find_after(exp_header, 'Rolling F3M Revenue')
    exp_profit_row = find_after(exp_header, 'Rolling F3M Profit')

    launch_manager = {
        'germany': {'revenue': gbb(ws, lm_rev_row, cols), 'profit_margin': gbb(ws, lm_profit_row, cols)},
        'pan_eu': {'revenue': gbb(ws, exp_rev_row, cols), 'profit_margin': gbb(ws, exp_profit_row, cols)},
    }

    # ---- Brand Manager sections ----
    section_re = re.compile(r'^BM\d+ - (.+?) \((PY1|Y1|Discontinued|Total)\)$')
    sections = []  # (row, brand, stage)
    for r in rows_sorted:
        m = section_re.match(labels[r])
        if m:
            sections.append((r, m.group(1).strip(), m.group(2)))

    brand_manager = {}
    for i, (r, brand, stage) in enumerate(sections):
        if stage not in ('PY1', 'Y1', 'Discontinued'):
            continue
        next_row = sections[i + 1][0] if i + 1 < len(sections) else (rows_sorted[-1] + 1)
        rev_row = margin_row = None
        for rr in rows_sorted:
            if r < rr < next_row:
                lbl = labels[rr].lower()
                if 'revenue' in lbl and 'growth' not in lbl and rev_row is None:
                    rev_row = rr
                if 'profit margin' in lbl and margin_row is None:
                    margin_row = rr
        stage_label = {'PY1': 'PY1', 'Y1': 'Y1 (F4-12)', 'Discontinued': 'Discontinued'}[stage]
        brand_manager.setdefault(brand, {})[stage_label] = {
            'revenue': gbb(ws, rev_row, cols) if rev_row else None,
            'profit_margin': gbb(ws, margin_row, cols) if margin_row else None,
        }

    return launch_manager, brand_manager


def extract_rd(ws, cols):
    """Leadership Scorecard 3, rows 13 onward: one row per named R&D product.
    Most labels are "Name - CODE" but a couple are "CODE - Name" -- handle both."""
    rd = {}
    r = 13
    while True:
        label = ws.cell(row=r, column=1).value
        if not label or str(label).strip().upper() == 'TOTAL':
            break
        label = str(label).strip()
        parts = [p.strip() for p in label.split('-', 1)]
        if len(parts) == 2:
            # Whichever side looks like a short alphanumeric code (no
            # spaces, has a digit or is short/uppercase-ish) is the code.
            a, b = parts
            def looks_like_code(s):
                return bool(re.match(r'^[A-Za-z0-9]+$', s)) and (any(ch.isdigit() for ch in s) or (s.isupper() and len(s) <= 6))
            if looks_like_code(a) and not looks_like_code(b):
                code = a
            elif looks_like_code(b):
                code = b
            else:
                code = b  # default: assume trailing segment is the code
        else:
            code = label
        rd[code] = {'label': label, 'revenue': gbb(ws, r, cols), 'profit_margin': None}  # margin intentionally blank
        r += 1
    return rd


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx_path')
    ap.add_argument('--month', required=True, help='YYYY-MM, e.g. 2026-08')
    ap.add_argument('--out', default='mapping/targets_monthly.json')
    args = ap.parse_args()

    month_num = int(args.month.split('-')[1])
    wb = load_workbook(args.xlsx_path, data_only=True)

    bm_ws = wb['BM Scorecard 3']
    bm_month_cols = find_month_columns(bm_ws)
    if month_num not in bm_month_cols:
        raise SystemExit(f'Month {month_num} not found in BM Scorecard 3 header row')
    launch_manager, brand_manager = extract_bm_scorecard(bm_ws, bm_month_cols[month_num])

    ls_ws = wb['Leadership Scorecard 3']
    ls_month_cols = find_month_columns(ls_ws)
    if month_num not in ls_month_cols:
        raise SystemExit(f'Month {month_num} not found in Leadership Scorecard 3 header row')
    rd_team = extract_rd(ls_ws, ls_month_cols[month_num])

    result = {
        'month': args.month,
        'note': 'Real per-month Good/Better/Best targets, sourced directly from BM Scorecard 3 (Launch Mgr, Brand Mgr) and Leadership Scorecard 3 (R&D, revenue only). Replaces the earlier quarterly/3 estimate for this month.',
        'rd_team': rd_team,
        'launch_manager': launch_manager,
        'brand_manager': brand_manager,
    }
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=1, ensure_ascii=False, default=str)
    print('Wrote', args.out)


if __name__ == '__main__':
    main()
